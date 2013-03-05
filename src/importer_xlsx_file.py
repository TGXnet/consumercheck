# StdLib imports
import os.path
# from StringIO import StringIO
from openpyxl.reader.excel import load_workbook
# import logging
# Log everything, and send it to stderr.
# http://docs.python.org/howto/logging-cookbook.html
# logging.basicConfig(level=logging.DEBUG)
# logging.basicConfig(level=logging.WARNING)


# Enthought imports
from traits.api import implements, HasTraits, File, Bool, Str, Int, List
from traitsui.api import View, Group, Item, TabularEditor, EnumEditor, Handler
from traitsui.tabular_adapter import TabularAdapter
from traitsui.menu import OKButton, CancelButton

# Local imports
from importer_interfaces import IDataImporter
from dataset import DS_TYPES, DataSet

#Import NumPy
import numpy as np


class RawLineAdapter(TabularAdapter):
    ncols = Int()

    #Temporary value to avoid crash
    columns = ['tmp']

    have_var_names = Bool(True)

    def _ncols_changed(self, info):
        self.columns = ["col{}".format(i) for i in range(self.ncols)]

preview_table = TabularEditor(
    adapter=RawLineAdapter(),
    operations=[],
    )

class FilePreviewer(Handler):
    _raw_lines = List(List)
    _parsed_data = List()

    def init(self, info):
        info.object.make_ds_name()
        self._probe_read(info.object)

    def object_have_var_names_changed(self, info):
        preview_table.adapter.have_var_names = info.object.have_var_names
        
#    def object_have_obj_names_changed(self, info):
#        preview_table.adapter.have_obj_names = info.object.have_obj_names

    def _probe_read(self, obj, no_lines=100, length=7):
        lines = []  
        raw_data = load_workbook(filename = obj.file_path)
        data_sheet = raw_data.get_active_sheet()
        
        if data_sheet.get_highest_row < no_lines:
            no_lines = data_sheet.get_highest_row
        if data_sheet.get_highest_column < length:
            length = data_sheet.get_highest_column

        preview_table.adapter.ncols = length

        c_table = []
        i=0
        j=0
        for row in data_sheet.range(data_sheet.calculate_dimension()):
            c_row = []
            
            for cell in row:
                c_row.append(cell.value)
                j+=1
                if j > length:
                    j=0
                    break
            c_table.append(c_row)
            i+=1
            if i > no_lines:
                break
        self._parsed_data = self._raw_lines = c_table
        


preview_handler = FilePreviewer()




class ImporterXlsxFile(HasTraits):
    implements(IDataImporter)
    file_path = File()
    transpose = Bool(False)
    have_var_names = Bool(True)
    have_obj_names = Bool(True)
    ds_id = Str()
    ds_name = Str()
    ds_type = Str()
    ds_type_list = List(DS_TYPES)

    def make_ds_name(self):
        # FIXME: Find a better more general solution
        fn = os.path.basename(self.file_path)
        fn = fn.partition('.')[0]
        fn = fn.lower()
        self.ds_id = self.ds_name = fn
        

    def import_data(self):
        self.ds = DataSet(
            ds_type=self.ds_type,
            display_name=self.ds_name,
            _source_file=self.file_path)
        
        raw_data = load_workbook(filename = self.file_path)
        data_sheet = raw_data.get_active_sheet()
        c_table = []
        c_col = []
        for row in data_sheet.range(data_sheet.calculate_dimension()):
            c_row = []
            i=0
            for cell in row:
                c_row.append(cell.value)
                if i==0:
                    i=1
                    c_col.append(cell.value)
            c_table.append(c_row)
        
        ###TODO: Col_values and row_values must be changed to something similar
        
        if self.have_obj_names:
            objnamelist = c_col
            if self.have_var_names:
                objnamelist.pop(0)
            
            for i in range(1,len(c_table)):
                c_table[i].pop(0)
            
            revised_list = []
            for sh in objnamelist:
                revised_list.append(unicode(sh))
            self.ds.object_names = revised_list
        
        if self.have_var_names:
            varnamelist = c_table[0]
            if self.have_obj_names:
                varnamelist.pop(0)
            c_table.pop(0)
            
            revised_list = []
            for sh in varnamelist:
                revised_list.append(unicode(sh))
            self.ds.variable_names = revised_list
        
        full_table = np.array(c_table)
        
        self.ds.matrix = full_table
        return self.ds


    pre_view = View(
        Group(
            Item('file_path', style='readonly'),
            Item('handler._parsed_data',
                 id='table',
                 editor=preview_table),
            ## Item('transpose'),
            Item('ds_id', style='readonly', label='File name'),
            Item('ds_name', label='Dataset name'),
            Item('ds_type', editor=EnumEditor(name='ds_type_list'), label='Dataset type'),
            Item('have_var_names', label='Existing variable names',
                 tooltip='Is first row variables names?'),
            Item('have_obj_names', label='Existing object names',
                 tooltip='Is first column object names?'),
            show_labels=True,
            ),
        title='Raw data preview',
        width=0.30,
        height=0.35,
        resizable=True,
        buttons=[CancelButton, OKButton],
        handler=preview_handler,
        kind='livemodal',
        )
    
# Run the demo (if invoked from the command line):
if __name__ == '__main__':
    test = ImporterXlsxFile()
    test.file_path = (os.path.join('datasets', 'Cheese', 'ConsumerLiking.xls'))
    test.configure_traits()
